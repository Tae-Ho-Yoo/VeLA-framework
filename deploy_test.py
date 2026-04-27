import os
import json
import re
import time
import logging
import traceback
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Union

import draccus
import json_numpy
import numpy as np
import pandas as pd
import torch
import uvicorn

from fastapi import FastAPI
from fastapi.responses import JSONResponse
from PIL import Image
from transformers import (
    AutoModelForVision2Seq,
    AutoProcessor,
    AutoTokenizer,
    AutoModelForCausalLM,
)
from openpyxl import Workbook

json_numpy.patch()

SYSTEM_PROMPT = (
    "A chat between a curious user and an artificial intelligence assistant. "
    "The assistant gives helpful, detailed, and polite answers to the user's questions."
)

FIXED_TASKS = [
    "extend the arm to reach for the cup",
    "reach out to take the cup",
    "extend arm to reach for the cup",
    "approach the cup",
]

FIXED_SPEEDS = [0.1, 0.2, 0.3, 0.4, 0.5]
LABELS = ["x", "y", "z", "roll", "pitch", "yaw", "gripper"]

DEFAULT_SPEED = 0.10
MIN_SPEED = 0.10
MAX_SPEED = 0.50

GLOBAL_COMFORT_MIN = 0.25
GLOBAL_COMFORT_MAX = 0.35
GLOBAL_COMFORT_CENTER = (GLOBAL_COMFORT_MIN + GLOBAL_COMFORT_MAX) / 2.0

ADAPT_GAIN = 0.35
MAX_DELTA_SPEED = 0.08

MIN_SAMPLES_FOR_PERSONAL_ZONE = 2
MAX_HISTORY_PER_USER = 20
MIN_PERSONAL_ZONE_WIDTH = 0.02

BASE_DIR = Path(__file__).resolve().parent
PROFILE_PATH = BASE_DIR / "user_speed_profiles.json"


# =========================================================
# 유틸
# =========================================================
def clamp_speed(speed: float) -> float:
    return max(MIN_SPEED, min(MAX_SPEED, float(speed)))


def clamp_zone(zone_min: float, zone_max: float) -> tuple[float, float]:
    zone_min = clamp_speed(zone_min)
    zone_max = clamp_speed(zone_max)
    if zone_min > zone_max:
        zone_min, zone_max = zone_max, zone_min
    return zone_min, zone_max


def ensure_min_zone_width(center: float, zone_min: float, zone_max: float) -> tuple[float, float]:
    width = zone_max - zone_min
    if width >= MIN_PERSONAL_ZONE_WIDTH:
        return clamp_zone(zone_min, zone_max)

    half = MIN_PERSONAL_ZONE_WIDTH / 2.0
    zone_min = center - half
    zone_max = center + half
    return clamp_zone(zone_min, zone_max)


def make_default_profile() -> dict[str, Any]:
    return {
        "comfortable_speeds": [],
        "current_speed": None,
    }


def load_user_speed_profiles() -> dict[str, dict[str, Any]]:
    if not PROFILE_PATH.exists():
        return {}

    try:
        with open(PROFILE_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)

        if not isinstance(data, dict):
            return {}

        cleaned: dict[str, dict[str, Any]] = {}

        for user_id, value in data.items():
            user_id = str(user_id)

            if isinstance(value, (int, float)):
                cleaned[user_id] = {
                    "comfortable_speeds": [clamp_speed(float(value))],
                    "current_speed": None,
                }
                continue

            if not isinstance(value, dict):
                cleaned[user_id] = make_default_profile()
                continue

            speeds = value.get("comfortable_speeds", [])
            if not isinstance(speeds, list):
                speeds = []

            safe_speeds = []
            for s in speeds:
                try:
                    safe_speeds.append(clamp_speed(float(s)))
                except (TypeError, ValueError):
                    continue

            current_speed = value.get("current_speed", None)
            try:
                current_speed = clamp_speed(float(current_speed)) if current_speed is not None else None
            except (TypeError, ValueError):
                current_speed = None

            cleaned[user_id] = {
                "comfortable_speeds": safe_speeds[-MAX_HISTORY_PER_USER:],
                "current_speed": current_speed,
            }

        return cleaned

    except Exception:
        logging.warning("[UserSpeedProfiles] Failed to load profile file. Starting with empty profiles.")
        return {}


def save_user_speed_profiles(profiles: dict[str, dict[str, Any]]) -> None:
    safe_profiles: dict[str, dict[str, Any]] = {}

    for user_id, profile in profiles.items():
        if not isinstance(profile, dict):
            continue

        speeds = profile.get("comfortable_speeds", [])
        if not isinstance(speeds, list):
            speeds = []

        safe_speeds = []
        for s in speeds[-MAX_HISTORY_PER_USER:]:
            try:
                safe_speeds.append(round(clamp_speed(float(s)), 4))
            except (TypeError, ValueError):
                continue

        current_speed = profile.get("current_speed", None)
        try:
            current_speed = round(clamp_speed(float(current_speed)), 4) if current_speed is not None else None
        except (TypeError, ValueError):
            current_speed = None

        safe_profiles[str(user_id)] = {
            "comfortable_speeds": safe_speeds,
            "current_speed": current_speed,
        }

    PROFILE_PATH.parent.mkdir(parents=True, exist_ok=True)

    with open(PROFILE_PATH, "w", encoding="utf-8") as f:
        json.dump(safe_profiles, f, indent=2, ensure_ascii=False)


def compute_personal_comfort_zone(
    user_id: str | None,
    profiles: dict[str, dict[str, Any]],
) -> tuple[float, float, float, bool, int]:
    if user_id is None or user_id not in profiles:
        return GLOBAL_COMFORT_MIN, GLOBAL_COMFORT_MAX, GLOBAL_COMFORT_CENTER, False, 0

    profile = profiles.get(user_id, {})
    speeds = profile.get("comfortable_speeds", [])
    if not isinstance(speeds, list):
        speeds = []

    safe_speeds = []
    for s in speeds:
        try:
            safe_speeds.append(clamp_speed(float(s)))
        except (TypeError, ValueError):
            continue

    num_samples = len(safe_speeds)

    if num_samples < MIN_SAMPLES_FOR_PERSONAL_ZONE:
        return GLOBAL_COMFORT_MIN, GLOBAL_COMFORT_MAX, GLOBAL_COMFORT_CENTER, False, num_samples

    zone_min = min(safe_speeds)
    zone_max = max(safe_speeds)
    center = (zone_min + zone_max) / 2.0

    zone_min, zone_max = ensure_min_zone_width(center, zone_min, zone_max)
    center = (zone_min + zone_max) / 2.0

    return zone_min, zone_max, center, True, num_samples


def get_base_speed(
    user_id: str | None,
    profiles: dict[str, dict[str, Any]],
) -> tuple[float, bool, str]:
    if user_id is None or user_id not in profiles:
        return DEFAULT_SPEED, False, "default"

    profile = profiles.get(user_id, {})

    current_speed = profile.get("current_speed", None)
    if current_speed is not None:
        try:
            return clamp_speed(float(current_speed)), True, "current_speed"
        except (TypeError, ValueError):
            pass

    speeds = profile.get("comfortable_speeds", [])
    if isinstance(speeds, list) and len(speeds) > 0:
        last_speed = speeds[-1]
        try:
            return clamp_speed(float(last_speed)), True, "preferred_speed"
        except (TypeError, ValueError):
            pass

    return DEFAULT_SPEED, False, "default"


def append_comfort_speed_to_profile(
    user_id: str | None,
    new_speed: float | None,
    profiles: dict[str, dict[str, Any]],
) -> bool:
    if user_id is None or new_speed is None:
        return False

    try:
        new_speed = clamp_speed(float(new_speed))
    except (TypeError, ValueError):
        return False

    if user_id not in profiles or not isinstance(profiles[user_id], dict):
        profiles[user_id] = make_default_profile()

    speeds = profiles[user_id].get("comfortable_speeds", [])
    if not isinstance(speeds, list):
        speeds = []

    speeds.append(new_speed)
    profiles[user_id]["comfortable_speeds"] = speeds[-MAX_HISTORY_PER_USER:]
    save_user_speed_profiles(profiles)
    return True


def sample_random_speed_outside_zone(zone_min: float, zone_max: float) -> float:
    zone_min, zone_max = clamp_zone(zone_min, zone_max)

    candidates = []
    if MIN_SPEED < zone_min:
        candidates.append((MIN_SPEED, zone_min))
    if zone_max < MAX_SPEED:
        candidates.append((zone_max, MAX_SPEED))

    if not candidates:
        return round(DEFAULT_SPEED, 4)

    lengths = [max(0.0, hi - lo) for lo, hi in candidates]
    total = sum(lengths)

    if total <= 1e-8:
        return round(DEFAULT_SPEED, 4)

    r = np.random.uniform(0.0, total)
    acc = 0.0

    for (lo, hi), seg_len in zip(candidates, lengths):
        if acc + seg_len >= r:
            return round(float(np.random.uniform(lo, hi)), 4)
        acc += seg_len

    lo, hi = candidates[-1]
    return round(float(np.random.uniform(lo, hi)), 4)


def get_openvla_prompt(instruction: str, openvla_path: Union[str, Path]) -> str:
    if "v01" in str(openvla_path):
        return (
            f"{SYSTEM_PROMPT} USER: What action should the robot take to "
            f"{instruction.lower()}? ASSISTANT:"
        )
    return f"In: What action should the robot take to {instruction.lower()}?\nOut:"


def get_speed_ratio(speed: float) -> float:
    denom = MAX_SPEED - MIN_SPEED
    if denom <= 0:
        return 0.5
    r = (float(speed) - MIN_SPEED) / denom
    return max(0.0, min(1.0, r))


def sanitize_single_step_task(task: str) -> str:
    if not task:
        return ""

    lines = task.splitlines()
    if not lines:
        return ""

    task = lines[0].strip()
    if not task:
        return ""

    lower_task = task.lower()

    separators = [" and ", " then ", ", then ", ";"]
    for sep in separators:
        if sep in lower_task:
            idx = lower_task.find(sep)
            task = task[:idx].strip()
            lower_task = task.lower()
            break

    for bad in [
        "Instruction:",
        "Reasoning:",
        "Task:",
        "Output:",
        "Refined Instruction:",
        "ComfortState:",
        "PaceState:",
    ]:
        bad_lower = bad.lower()
        task_lower = task.lower()
        if bad_lower in task_lower:
            idx = task_lower.find(bad_lower)
            task = task[:idx].strip()

    speed_words = ["slow", "slowly", "fast", "quickly", "gently", "rapidly"]
    for word in speed_words:
        task = re.sub(rf"\b{word}\b", "", task, flags=re.IGNORECASE)

    task = re.sub(r"\s+", " ", task).strip(" \n\r\t:.-\"'")
    return task


def shorten_reasoning(reasoning: str) -> str:
    reasoning = reasoning.strip()
    if "." in reasoning:
        reasoning = reasoning.split(".")[0].strip() + "."
    return reasoning


def mae7(action: list[float], base: list[float]) -> float:
    a = np.array(action[:7], dtype=float)
    b = np.array(base[:7], dtype=float)
    return float(np.mean(np.abs(a - b)))


def mae6(a: list[float], b: list[float]) -> float:
    return float(np.mean(np.abs(np.array(a[:6]) - np.array(b[:6]))))


def format_float_list(values: list[float], precision: int = 6) -> str:
    return "[" + ", ".join(f"{float(v):.{precision}f}" for v in values) + "]"


def print_fixed_task_action_table(results: list[dict[str, Any]]) -> None:
    if not results:
        print("\n[INFO] No results to display.")
        return

    task_w = max(len(str(r.get("task", ""))) for r in results) + 2
    total_w = task_w + 10 + 11 * 7

    print("\n" + "=" * total_w)
    print("Fixed Task × Speed -> 7D Action")
    print("=" * total_w)

    header = (
        f"{'task':<{task_w}}"
        f"{'speed':>8}  "
        f"{'x':>10}  {'y':>10}  {'z':>10}  "
        f"{'roll':>10}  {'pitch':>10}  {'yaw':>10}  {'gripper':>10}"
    )
    print(header)
    print("-" * total_w)

    for r in results:
        task = r.get("task", "")
        speed = r.get("input_speed", "")
        action = r.get("action", None)

        if action is None:
            print(f"{task:<{task_w}}{str(speed):>8}  ERROR")
            continue

        print(
            f"{task:<{task_w}}"
            f"{float(speed):>8.1f}  "
            f"{float(action[0]):>10.6f}  {float(action[1]):>10.6f}  {float(action[2]):>10.6f}  "
            f"{float(action[3]):>10.6f}  {float(action[4]):>10.6f}  {float(action[5]):>10.6f}  {float(action[6]):>10.6f}"
        )


def save_fixed_task_action_results_json(results: list[dict[str, Any]], output_path: str | Path) -> None:
    serializable = []
    for r in results:
        serializable.append({
            "task": r.get("task"),
            "input_speed": r.get("input_speed"),
            "speed_scale": r.get("speed_scale"),
            "raw_action": r.get("raw_action"),
            "action": r.get("action"),
            "error": r.get("error"),
        })

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(serializable, f, indent=2, ensure_ascii=False)


def parse_reasoning_task(generated_text: str) -> Dict[str, str]:
    reasoning_match = re.search(r"Reasoning:\s*(.*)", generated_text)
    task_match = re.search(r"(?:Task|Refined Instruction):\s*(.*)", generated_text)
    pace_match = re.search(r"PaceState:\s*(.*)", generated_text)
    comfort_match = re.search(r"ComfortState:\s*(.*)", generated_text)

    reasoning = reasoning_match.group(1).strip() if reasoning_match else ""
    task = task_match.group(1).strip() if task_match else ""
    pace_state = pace_match.group(1).strip().lower() if pace_match else ""
    comfort_state = comfort_match.group(1).strip().lower() if comfort_match else ""

    task = sanitize_single_step_task(task)
    reasoning = shorten_reasoning(reasoning) if reasoning else ""

    valid_pace = {"too_fast", "comfortable", "too_slow"}
    if pace_state not in valid_pace:
        pace_state = "comfortable"

    valid_comfort = {"comfortable", "uncomfortable", "neutral"}
    if comfort_state not in valid_comfort:
        comfort_state = "neutral"

    return {
        "reasoning": reasoning,
        "task": task,
        "pace_state": pace_state,
        "comfort_state": comfort_state,
    }


def compute_speed_update_continuous(
    pace_state: str,
    emotion_history: list[str] | None,
    base_speed: float | None,
    comfort_center: float,
):
    if base_speed is None:
        return 1.0, base_speed, None, 0.0

    try:
        base_speed = float(base_speed)
    except (TypeError, ValueError):
        return 1.0, None, None, 0.0

    speed_ratio = get_speed_ratio(base_speed)

    latest_emotion = None
    if isinstance(emotion_history, list) and len(emotion_history) > 0:
        latest_emotion = str(emotion_history[-1]).lower().strip()

    target_speed = base_speed
    delta_speed = 0.0

    if latest_emotion == "negative":
        desired_speed = comfort_center
        error = desired_speed - base_speed
        raw_delta = ADAPT_GAIN * error
        delta_speed = max(-MAX_DELTA_SPEED, min(MAX_DELTA_SPEED, raw_delta))
        target_speed = round(clamp_speed(base_speed + delta_speed), 4)

    if base_speed <= 1e-6:
        scale = 1.0
    else:
        scale = target_speed / base_speed

    return scale, target_speed, speed_ratio, delta_speed


def apply_speed_modulation(
    action,
    pace_state: str,
    emotion_history: list[str] | None,
    base_speed: float | None = None,
    comfort_center: float = GLOBAL_COMFORT_CENTER,
):
    if action is None:
        return action, 1.0, base_speed, None, 0.0

    scale, target_speed, speed_ratio, delta_speed = compute_speed_update_continuous(
        pace_state,
        emotion_history,
        base_speed,
        comfort_center,
    )

    modulated = list(action)
    for i in range(min(6, len(modulated))):
        modulated[i] *= scale

    return modulated, scale, target_speed, speed_ratio, delta_speed


EMOTION_RESULTS = [
    {
        "emotion": "positive",
        "speed": 0.1,
        "task_used": "Extend the arm to reach for the cup",
        "speed_scale": 1.0,
        "target_speed": 0.1,
        "action": [0.002475303572766905, 0.006698593539350123, 0.005994562820038369,
                   -0.01929670767340016, -0.010828568453297924, 0.04184003857654672, 0.996078431372549],
    },
    {
        "emotion": "positive",
        "speed": 0.2,
        "task_used": "Reach out to take the cup",
        "speed_scale": 1.0,
        "target_speed": 0.2,
        "action": [-0.006695340317838347, -0.002690441256060362, -0.011673555954852527,
                   -0.009717794240689859, 0.03137290809610306, -0.011349297856583396, 0.996078431372549],
    },
    {
        "emotion": "positive",
        "speed": 0.3,
        "task_used": "Reach out to take the cup",
        "speed_scale": 1.0,
        "target_speed": 0.3,
        "action": [-0.006695340317838347, -0.002690441256060362, -0.011673555954852527,
                   -0.009717794240689859, 0.03137290809610306, -0.011349297856583396, 0.996078431372549],
    },
    {
        "emotion": "positive",
        "speed": 0.4,
        "task_used": "Reach out to take the cup",
        "speed_scale": 1.0,
        "target_speed": 0.4,
        "action": [-0.006695340317838347, -0.002690441256060362, -0.011673555954852527,
                   -0.009717794240689859, 0.03137290809610306, -0.011349297856583396, 0.996078431372549],
    },
    {
        "emotion": "positive",
        "speed": 0.5,
        "task_used": "Reach out to take the cup",
        "speed_scale": 1.0,
        "target_speed": 0.5,
        "action": [-0.006695340317838347, -0.002690441256060362, -0.011673555954852527,
                   -0.009717794240689859, 0.03137290809610306, -0.011349297856583396, 0.996078431372549],
    },

    {
        "emotion": "negative",
        "speed": 0.1,
        "task_used": "Approach the cup",
        "speed_scale": 1.7,
        "target_speed": 0.17,
        "action": [-0.01100183233022692, -0.019434325897693686, 0.04420188543573019,
                   -0.017605860398213255, -0.037767656406760254, -0.03847420343359326, 0.996078431372549],
    },
    {
        "emotion": "negative",
        "speed": 0.2,
        "task_used": "Approach the cup",
        "speed_scale": 1.175,
        "target_speed": 0.235,
        "action": [-0.007604207640009782, -0.013432548782229459, 0.030551303168813514,
                   -0.012168756451706219, -0.026104115457613705, -0.026592464137924748, 0.996078431372549],
    },
    {
        "emotion": "negative",
        "speed": 0.3,
        "task_used": "Extend arm to reach for the cup",
        "speed_scale": 1.0,
        "target_speed": 0.3,
        "action": [0.002475303572766905, -0.010784436769345263, -0.011153905402649857,
                   -0.0033318519522163442, 0.017975613953436095, -0.004902105561658515, 0.996078431372549],
    },
    {
        "emotion": "negative",
        "speed": 0.4,
        "task_used": "Extend arm to reach for the cup",
        "speed_scale": 0.9125,
        "target_speed": 0.365,
        "action": [0.0022587145101498006, -0.009840798552027552, -0.010177938679917994,
                   -0.003040314906397414, 0.016402747732510436, -0.004473171325013395, 0.996078431372549],
    },
    {
        "emotion": "negative",
        "speed": 0.5,
        "task_used": "Extend arm to reach for the cup",
        "speed_scale": 0.86,
        "target_speed": 0.43,
        "action": [0.002128761072579538, -0.009274615621636926, -0.009592358646278876,
                   -0.002865392678906056, 0.015459027999955043, -0.004215810783026323, 0.996078431372549],
    },
]


def recover_raw_action(action: list[float], speed_scale: float | int | None) -> list[float]:
    if not isinstance(action, list) or len(action) != 7:
        raise ValueError(f"action must be a list of length 7, got: {action}")

    if speed_scale is None:
        scale = 1.0
    else:
        try:
            scale = float(speed_scale)
        except (TypeError, ValueError):
            scale = 1.0

    if abs(scale) < 1e-12:
        scale = 1.0

    recovered = []
    for i in range(6):
        recovered.append(float(action[i]) / scale)
    recovered.append(float(action[6]))
    return recovered


def write_compare_section(
    ws,
    start_row: int,
    title: str,
    emotion: str,
    speed: float,
    task_used: str,
    final_action: list[float],
    recovered_action: list[float],
    compare_targets: list[tuple[str, list[float]]],
    speed_scale: float | int | None,
    target_speed: float | str | None,
) -> int:
    ws.cell(row=start_row, column=1, value=title)
    start_row += 1

    header = ["emotion", "speed", "task_used", "base_task", "speed_scale", "target_speed"]
    header += [f"final_{label}" for label in LABELS]
    header += [f"recovered_{label}" for label in LABELS]
    header += [f"base_{label}" for label in LABELS]
    header += [f"{label}_err" for label in LABELS]
    header += ["l1_7d"]

    for col_idx, value in enumerate(header, start=1):
        ws.cell(row=start_row, column=col_idx, value=value)
    start_row += 1

    for base_task_name, base_action in compare_targets:
        errs = [float(recovered_action[i] - base_action[i]) for i in range(7)]

        row = [
            emotion,
            speed,
            task_used,
            base_task_name,
            round(float(speed_scale), 6) if speed_scale is not None else "",
            target_speed,
        ]
        row += [round(float(final_action[i]), 6) for i in range(7)]
        row += [round(float(recovered_action[i]), 6) for i in range(7)]
        row += [round(float(base_action[i]), 6) for i in range(7)]
        row += [round(float(errs[i]), 6) for i in range(7)]
        row += [round(mae7(recovered_action, base_action), 6)]

        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=start_row, column=col_idx, value=value)
        start_row += 1

    return start_row + 1


def save_emotion_summary_excel(
    base_actions: dict[str, list[float]],
    single_base_action: list[float],
    emotion_results: list[dict[str, Any]],
    output_path: str | Path,
    single_base_task_name: str = "move toward the cup",
) -> None:
    output_path = Path(output_path)

    wb = Workbook()
    default_ws = wb.active
    default_ws.title = "README"
    default_ws.append(["This workbook has 2 sheets: positive_compare and negative_compare"])
    default_ws.append(["Each sheet contains all speeds 0.1~0.5"])
    default_ws.append(["For each speed, two comparison sections are included:"])
    default_ws.append(["1) original_4task_compare"])
    default_ws.append(["2) single_base_compare"])

    for emotion in ["positive", "negative"]:
        ws = wb.create_sheet(title=f"{emotion}_compare")
        row_ptr = 1

        for speed in FIXED_SPEEDS:
            matched = [
                item for item in emotion_results
                if item["emotion"] == emotion and float(item["speed"]) == float(speed)
            ]

            if not matched:
                ws.cell(row=row_ptr, column=1, value=f"{emotion} speed={speed}")
                ws.cell(row=row_ptr + 1, column=1, value="No result provided")
                row_ptr += 4
                continue

            item = matched[0]
            task_used = item["task_used"]
            final_action = item["action"]
            speed_scale = item.get("speed_scale", 1.0)
            target_speed = item.get("target_speed", "")
            recovered_action = recover_raw_action(final_action, speed_scale)

            ws.cell(row=row_ptr, column=1, value=f"{emotion} speed={speed}")
            row_ptr += 1

            compare_targets_4task = [(task_name, base_actions[task_name]) for task_name in FIXED_TASKS]
            row_ptr = write_compare_section(
                ws=ws,
                start_row=row_ptr,
                title="original_4task_compare",
                emotion=emotion,
                speed=speed,
                task_used=task_used,
                final_action=final_action,
                recovered_action=recovered_action,
                compare_targets=compare_targets_4task,
                speed_scale=speed_scale,
                target_speed=target_speed,
            )

            compare_targets_single = [(single_base_task_name, single_base_action)]
            row_ptr = write_compare_section(
                ws=ws,
                start_row=row_ptr,
                title="single_base_compare",
                emotion=emotion,
                speed=speed,
                task_used=task_used,
                final_action=final_action,
                recovered_action=recovered_action,
                compare_targets=compare_targets_single,
                speed_scale=speed_scale,
                target_speed=target_speed,
            )

            row_ptr += 1

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)
    print(f"\nSaved extra Excel: {output_path}")


class OpenVLAServer:
    def __init__(
        self,
        openvla_path: Union[str, Path],
        reasoning_model_name: str = "Qwen/Qwen2-7B-Instruct",
    ) -> None:
        self.openvla_path = openvla_path
        self.reasoning_model_name = reasoning_model_name

        self.device = torch.device("cuda:0") if torch.cuda.is_available() else torch.device("cpu")
        self.reasoning_device = torch.device("cpu")

        self.processor = AutoProcessor.from_pretrained(
            self.openvla_path,
            trust_remote_code=True,
        )

        self.vla = AutoModelForVision2Seq.from_pretrained(
            self.openvla_path,
            torch_dtype=torch.bfloat16 if self.device.type == "cuda" else torch.float32,
            low_cpu_mem_usage=True,
            trust_remote_code=True,
        ).to(self.device)
        self.vla.eval()

        self.reasoning_tokenizer = AutoTokenizer.from_pretrained(self.reasoning_model_name)
        self.reasoning_model = AutoModelForCausalLM.from_pretrained(
            self.reasoning_model_name,
            torch_dtype=torch.float32,
            low_cpu_mem_usage=True,
        ).to(self.reasoning_device)
        self.reasoning_model.eval()

        if os.path.isdir(self.openvla_path):
            stats_path = Path(self.openvla_path) / "dataset_statistics.json"
            if stats_path.exists():
                with open(stats_path, "r", encoding="utf-8") as f:
                    self.vla.norm_stats = json.load(f)

        self.user_speed_profiles = load_user_speed_profiles()

        logging.warning(f"[OpenVLA device] {self.device}")
        logging.warning(f"[Reasoner device] {self.reasoning_device}")
        logging.warning(f"[Reasoner model] {self.reasoning_model_name}")
        logging.warning(
            f"[Global comfort zone] {GLOBAL_COMFORT_MIN:.2f} ~ {GLOBAL_COMFORT_MAX:.2f} "
            f"(center={GLOBAL_COMFORT_CENTER:.2f}) m/s"
        )
        logging.warning(f"[Loaded user speed profiles] {self.user_speed_profiles}")

        if hasattr(self.vla, "norm_stats"):
            logging.warning(f"[Norm stats keys] {list(self.vla.norm_stats.keys())}")

    def refine_instruction(
        self,
        instruction: str,
        emotion_history: list[str],
        current_speed: float | None,
    ) -> Dict[str, str]:
        prompt = get_reasoner_prompt(instruction, emotion_history, current_speed)

        inputs = self.reasoning_tokenizer(prompt, return_tensors="pt")
        inputs = {k: v.to(self.reasoning_device) for k, v in inputs.items()}

        with torch.no_grad():
            outputs = self.reasoning_model.generate(
                **inputs,
                max_new_tokens=96,
                do_sample=False,
                pad_token_id=self.reasoning_tokenizer.eos_token_id,
            )

        generated_ids = outputs[0][inputs["input_ids"].shape[1]:]
        generated_text = self.reasoning_tokenizer.decode(
            generated_ids,
            skip_special_tokens=True,
        ).strip()

        if not generated_text.startswith("Reasoning:"):
            generated_text = "Reasoning: " + generated_text

        print("[RAW_REASONER_OUTPUT]")
        print(generated_text)

        return parse_reasoning_task(generated_text)

    def predict_action_from_fixed_task(
        self,
        image: np.ndarray,
        task: str,
        speed: float,
        unnorm_key: str = "bridge_orig",
    ) -> dict[str, Any]:
        try:
            speed = clamp_speed(speed)
            task = sanitize_single_step_task(task)
            prompt = get_openvla_prompt(task, self.openvla_path)

            if self.device.type == "cuda":
                inputs = self.processor(
                    prompt,
                    Image.fromarray(image).convert("RGB"),
                ).to(self.device, dtype=torch.bfloat16)
            else:
                inputs = self.processor(
                    prompt,
                    Image.fromarray(image).convert("RGB"),
                ).to(self.device, dtype=torch.float32)

            if self.device.type == "cuda":
                torch.cuda.synchronize()

            start = time.time()
            action = self.vla.predict_action(
                **inputs,
                unnorm_key=unnorm_key,
                do_sample=False,
            )
            if self.device.type == "cuda":
                torch.cuda.synchronize()
            end = time.time()

            print(f"[FIXED_EVAL] task='{task}' | speed={speed:.1f} | infer={end - start:.3f}s")

            if hasattr(action, "tolist"):
                action = action.tolist()

            raw_action = list(action)
            modulated_action = list(action)

            base_scale = speed / DEFAULT_SPEED if DEFAULT_SPEED > 1e-6 else 1.0
            for i in range(min(6, len(modulated_action))):
                modulated_action[i] *= base_scale

            return {
                "task": task,
                "input_speed": speed,
                "speed_scale": round(base_scale, 4),
                "raw_action": [round(float(x), 6) for x in raw_action],
                "action": [round(float(x), 6) for x in modulated_action],
                "error": None,
            }

        except Exception as e:
            return {
                "task": task,
                "input_speed": speed,
                "speed_scale": None,
                "raw_action": None,
                "action": None,
                "error": str(e),
            }

    def generate_fixed_task_action_table(
        self,
        image: np.ndarray,
        unnorm_key: str = "bridge_orig",
    ) -> list[dict[str, Any]]:
        results = []
        for task in FIXED_TASKS:
            for speed in FIXED_SPEEDS:
                out = self.predict_action_from_fixed_task(
                    image=image,
                    task=task,
                    speed=speed,
                    unnorm_key=unnorm_key,
                )
                results.append(out)
        return results

    def generate_openvla_base_actions(
        self,
        image: np.ndarray,
        unnorm_key: str = "bridge_orig",
    ) -> dict[str, list[float]]:
        base_actions: dict[str, list[float]] = {}
        for task in FIXED_TASKS:
            out = self.predict_action_from_fixed_task(
                image=image,
                task=task,
                speed=DEFAULT_SPEED,
                unnorm_key=unnorm_key,
            )
            base_actions[task] = out["raw_action"]
        return base_actions
    
    def generate_emotion_speed_results(
        self,
        image: np.ndarray,
        emotion_label: str,
        unnorm_key: str = "bridge_orig",
    ) -> list[dict[str, Any]]:
        results = []
        for task in FIXED_TASKS:
            for speed in FIXED_SPEEDS:
                out = self.predict_action_from_fixed_task(
                    image=image,
                    task=task,
                    speed=speed,
                    unnorm_key=unnorm_key,
                )
                out["emotion"] = emotion_label
                results.append(out)
        return results    

    def generate_single_base_action(
        self,
        image: np.ndarray,
        base_task: str = "move toward the cup",
        unnorm_key: str = "bridge_orig",
    ) -> list[float]:
        out = self.predict_action_from_fixed_task(
            image=image,
            task=base_task,
            speed=DEFAULT_SPEED,
            unnorm_key=unnorm_key,
        )
        if out["raw_action"] is None:
            raise ValueError(f"Failed to generate single base action for task: {base_task}")
        return out["raw_action"]

    def predict_action(self, payload: Dict[str, Any]) -> JSONResponse:
        try:
            if double_encode := ("encoded" in payload):
                assert len(payload.keys()) == 1
                payload = json.loads(payload["encoded"])

            image = payload.get("image", None)
            if image is None:
                fallback_path = Path("test.jpg")
                if not fallback_path.exists():
                    raise FileNotFoundError("No image provided and fallback test.jpg not found.")
                image = Image.open(fallback_path).convert("RGB").resize((256, 256))
                image = np.array(image)

            instruction = payload["instruction"]
            user_id = payload.get("user_id", None)

            emotion_history = payload.get("emotion_history", [])
            if not isinstance(emotion_history, list):
                emotion_history = []

            unnorm_key = payload.get("unnorm_key", "bridge_orig")

            comfort_min, comfort_max, comfort_center, used_personal_zone, num_samples = (
                compute_personal_comfort_zone(user_id, self.user_speed_profiles)
            )

            base_speed, has_preferred_speed, base_speed_source = get_base_speed(
                user_id,
                self.user_speed_profiles,
            )

            reasoner_result = self.refine_instruction(
                instruction,
                emotion_history,
                base_speed,
            )

            refined_instruction = reasoner_result["task"]
            pace_state = reasoner_result["pace_state"]
            comfort_state = reasoner_result["comfort_state"]

            if not reasoner_result["reasoning"] or not refined_instruction:
                result = {
                    "instruction": instruction,
                    "user_id": user_id,
                    "emotion_history": emotion_history,
                    "preferred_speed_found": has_preferred_speed,
                    "base_speed": base_speed,
                    "base_speed_source": base_speed_source,
                    "reasoning": reasoner_result["reasoning"],
                    "task": refined_instruction,
                    "pace_state": pace_state,
                    "comfort_state": comfort_state,
                    "comfort_zone_min": comfort_min,
                    "comfort_zone_max": comfort_max,
                    "comfort_zone_center": comfort_center,
                    "personalized_comfort_zone": used_personal_zone,
                    "comfort_zone_samples": num_samples,
                    "speed_scale": 1.0,
                    "target_speed": base_speed,
                    "profile_saved": False,
                    "saved_speed": None,
                    "saved_speeds": self.user_speed_profiles.get(user_id, {}).get("comfortable_speeds", []) if user_id else [],
                    "current_speed_next": self.user_speed_profiles.get(user_id, {}).get("current_speed", None) if user_id else None,
                    "speed_ratio": None,
                    "speed_delta": 0.0,
                    "action": None,
                }
                if double_encode:
                    return JSONResponse(json_numpy.dumps(result))
                return JSONResponse(result)

            prompt = get_openvla_prompt(refined_instruction, self.openvla_path)

            if self.device.type == "cuda":
                inputs = self.processor(
                    prompt,
                    Image.fromarray(image).convert("RGB"),
                ).to(self.device, dtype=torch.bfloat16)
            else:
                inputs = self.processor(
                    prompt,
                    Image.fromarray(image).convert("RGB"),
                ).to(self.device, dtype=torch.float32)

            if self.device.type == "cuda":
                torch.cuda.synchronize()

            action = self.vla.predict_action(
                **inputs,
                unnorm_key=unnorm_key,
                do_sample=False,
            )

            if self.device.type == "cuda":
                torch.cuda.synchronize()

            if hasattr(action, "tolist"):
                action = action.tolist()

            modulated_action, speed_scale, target_speed, speed_ratio, speed_delta = (
                apply_speed_modulation(
                    action,
                    pace_state,
                    emotion_history,
                    base_speed,
                    comfort_center,
                )
            )

            profile_saved = False
            saved_speed = None

            latest_emotion = None
            if isinstance(emotion_history, list) and len(emotion_history) > 0:
                latest_emotion = str(emotion_history[-1]).lower().strip()

            if user_id is not None:
                if user_id not in self.user_speed_profiles or not isinstance(self.user_speed_profiles[user_id], dict):
                    self.user_speed_profiles[user_id] = make_default_profile()

            if user_id is not None and target_speed is not None and latest_emotion == "positive":
                profile_saved = append_comfort_speed_to_profile(
                    user_id=user_id,
                    new_speed=target_speed,
                    profiles=self.user_speed_profiles,
                )
                if profile_saved:
                    saved_speeds = self.user_speed_profiles[user_id]["comfortable_speeds"]
                    saved_speed = saved_speeds[-1] if saved_speeds else None

                self.user_speed_profiles[user_id]["current_speed"] = sample_random_speed_outside_zone(
                    comfort_min,
                    comfort_max,
                )
                save_user_speed_profiles(self.user_speed_profiles)

            elif user_id is not None and target_speed is not None and latest_emotion == "negative":
                self.user_speed_profiles[user_id]["current_speed"] = clamp_speed(target_speed)
                save_user_speed_profiles(self.user_speed_profiles)

            elif user_id is not None and target_speed is not None:
                self.user_speed_profiles[user_id]["current_speed"] = clamp_speed(target_speed)
                save_user_speed_profiles(self.user_speed_profiles)

            saved_speeds = self.user_speed_profiles.get(user_id, {}).get("comfortable_speeds", []) if user_id else []
            current_speed_next = self.user_speed_profiles.get(user_id, {}).get("current_speed", None) if user_id else None

            result = {
                "instruction": instruction,
                "user_id": user_id,
                "emotion_history": emotion_history,
                "preferred_speed_found": has_preferred_speed,
                "base_speed": base_speed,
                "base_speed_source": base_speed_source,
                "reasoning": reasoner_result["reasoning"],
                "task": refined_instruction,
                "pace_state": pace_state,
                "comfort_state": comfort_state,
                "comfort_zone_min": comfort_min,
                "comfort_zone_max": comfort_max,
                "comfort_zone_center": comfort_center,
                "personalized_comfort_zone": used_personal_zone,
                "comfort_zone_samples": num_samples,
                "speed_scale": speed_scale,
                "target_speed": target_speed,
                "profile_saved": profile_saved,
                "saved_speed": saved_speed,
                "saved_speeds": saved_speeds,
                "current_speed_next": current_speed_next,
                "speed_ratio": speed_ratio,
                "speed_delta": speed_delta,
                "action": modulated_action,
            }

            if double_encode:
                return JSONResponse(json_numpy.dumps(result))
            return JSONResponse(result)

        except Exception:
            logging.error(traceback.format_exc())
            return JSONResponse({"error": "generation failed"}, status_code=500)

    def run(self, host: str = "0.0.0.0", port: int = 8000) -> None:
        self.app = FastAPI()
        self.app.post("/act")(self.predict_action)
        uvicorn.run(self.app, host=host, port=port)


@dataclass
class DeployConfig:
    openvla_path: Union[str, Path] = "openvla/openvla-7b"
    reasoning_model_name: str = "Qwen/Qwen2-7B-Instruct"

    mode: str = "server"
    host: str = "0.0.0.0"
    port: int = 8000

    eval_image_path: str = "test.jpg"
    eval_unnorm_key: str = "bridge_orig"

    eval_save_json: bool = True
    eval_output_json_path: str = "fixed_task_actions.json"

    eval_save_compare_json: bool = True
    eval_compare_json_path: str = "openvla_vs_emotion_compare.json"

    eval_save_speed_task_compare_json: bool = True
    eval_speed_task_compare_json_path: str = "speed_task_compare_tables.json"

    eval_generate_compare_actions_py: bool = False
    eval_compare_actions_py_path: str = "compare_actions_generated.py"

    eval_save_excel: bool = True
    eval_excel_path: str = "emotion_vs_base_tasks.xlsx"

    # 추가 출력
    eval_save_extra_emotion_summary_excel: bool = True
    eval_extra_emotion_summary_excel_path: str = "emotion_compare_summary.xlsx"
    eval_single_base_task: str = "approach the cup"


@draccus.wrap()
def deploy(cfg: DeployConfig) -> None:
    server = OpenVLAServer(
        openvla_path=cfg.openvla_path,
        reasoning_model_name=cfg.reasoning_model_name,
    )

    if cfg.mode == "server":
        server.run(cfg.host, port=cfg.port)
        return

    if cfg.mode == "fixed_eval":
        image_path = Path(cfg.eval_image_path)
        if not image_path.exists():
            raise FileNotFoundError(f"Evaluation image not found: {image_path}")

        image = Image.open(image_path).convert("RGB").resize((256, 256))
        image = np.array(image)

        results = server.generate_fixed_task_action_table(
            image=image,
            unnorm_key=cfg.eval_unnorm_key,
        )

        print_fixed_task_action_table(results)

        if cfg.eval_save_json:
            save_fixed_task_action_results_json(results, cfg.eval_output_json_path)
            print(f"\nSaved JSON: {cfg.eval_output_json_path}")

        base_actions = server.generate_openvla_base_actions(
            image=image,
            unnorm_key=cfg.eval_unnorm_key,
        )

        print("\n" + "=" * 120)
        print("OpenVLA Base Actions (per task, speed=0.1 raw reference)")
        print("=" * 120)
        for task, action in base_actions.items():
            print(task)
            print(action)
            print()

        pos_results = server.generate_emotion_speed_results(
            image=image,
            emotion_label="positive",
            unnorm_key=cfg.eval_unnorm_key,
        )
        neg_results = server.generate_emotion_speed_results(
            image=image,
            emotion_label="negative",
            unnorm_key=cfg.eval_unnorm_key,
        )

        all_compare_results = pos_results + neg_results

        if cfg.eval_save_compare_json:
            rows = []
            for r in all_compare_results:
                task = r["task"]
                action = r["action"]

                if action is None:
                    rows.append({
                        "emotion": r.get("emotion"),
                        "task": task,
                        "input_speed": r["input_speed"],
                        "base_action": base_actions.get(task),
                        "action": None,
                        "error": r.get("error"),
                    })
                    continue

                base = base_actions[task]
                rows.append({
                    "emotion": r.get("emotion"),
                    "task": task,
                    "input_speed": r["input_speed"],
                    "base_action": base,
                    "action": action,
                    "diff": {
                        "x": round(float(action[0] - base[0]), 6),
                        "y": round(float(action[1] - base[1]), 6),
                        "z": round(float(action[2] - base[2]), 6),
                        "roll": round(float(action[3] - base[3]), 6),
                        "pitch": round(float(action[4] - base[4]), 6),
                        "yaw": round(float(action[5] - base[5]), 6),
                        "gripper": round(float(action[6] - base[6]), 6),
                        "l1_6d": round(float(mae6(action, base)), 6),
                        "l1_7d": round(float(mae7(action, base)), 6),
                    },
                    "error": None,
                })

            with open(cfg.eval_compare_json_path, "w", encoding="utf-8") as f:
                json.dump(rows, f, indent=2, ensure_ascii=False)
            print(f"\nSaved Compare JSON: {cfg.eval_compare_json_path}")

        if cfg.eval_save_speed_task_compare_json:
            speed_task_compare_serializable = []
            for emotion in ["positive", "negative"]:
                filtered = [r for r in all_compare_results if str(r.get("emotion", "")).lower() == emotion]
                for speed in FIXED_SPEEDS:
                    rows = sorted(
                        [r for r in filtered if float(r["input_speed"]) == float(speed)],
                        key=lambda x: FIXED_TASKS.index(x["task"]),
                    )

                    table_rows = []
                    for r in rows:
                        task = r["task"]
                        action = r["action"]
                        base = base_actions[task]

                        if action is None:
                            table_rows.append({
                                "task": task,
                                "base_action": base,
                                "action": None,
                                "diff": None,
                                "l1_6d": None,
                                "l1_7d": None,
                                "error": r.get("error"),
                            })
                        else:
                            table_rows.append({
                                "task": task,
                                "base_action": base,
                                "action": action,
                                "diff": {
                                    "x": round(float(action[0] - base[0]), 6),
                                    "y": round(float(action[1] - base[1]), 6),
                                    "z": round(float(action[2] - base[2]), 6),
                                    "roll": round(float(action[3] - base[3]), 6),
                                    "pitch": round(float(action[4] - base[4]), 6),
                                    "yaw": round(float(action[5] - base[5]), 6),
                                    "gripper": round(float(action[6] - base[6]), 6),
                                },
                                "l1_6d": round(float(mae6(action, base)), 6),
                                "l1_7d": round(float(mae7(action, base)), 6),
                                "error": None,
                            })

                    speed_task_compare_serializable.append({
                        "emotion": emotion,
                        "speed": speed,
                        "rows": table_rows,
                    })

            with open(cfg.eval_speed_task_compare_json_path, "w", encoding="utf-8") as f:
                json.dump(speed_task_compare_serializable, f, indent=2, ensure_ascii=False)
            print(f"\nSaved Speed-Task Compare JSON: {cfg.eval_speed_task_compare_json_path}")

        # 원래 출력 유지
        if cfg.eval_save_excel:
            # 기존 함수가 있다면 그대로 유지하고 싶다는 뜻으로 남겨둠
            # 필요 없으면 이 블록은 삭제 가능
            pass

        # 추가 출력
        if cfg.eval_save_extra_emotion_summary_excel:
            single_base_action = server.generate_single_base_action(
                image=image,
                base_task=cfg.eval_single_base_task,
                unnorm_key=cfg.eval_unnorm_key,
            )

            save_emotion_summary_excel(
                base_actions=base_actions,
                single_base_action=single_base_action,
                emotion_results=EMOTION_RESULTS,
                output_path=cfg.eval_extra_emotion_summary_excel_path,
                single_base_task_name=cfg.eval_single_base_task,
            )

        return

    raise ValueError(f"Unsupported mode: {cfg.mode}")


if __name__ == "__main__":
    deploy()